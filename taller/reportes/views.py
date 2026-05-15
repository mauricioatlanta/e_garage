# -----------------------------------------------------------------------------
# Copyright (c) 2025 eGarage. Todos los derechos reservados.
#
# PROPIEDAD INTELECTUAL PROTEGIDA. ESTRICTAMENTE CONFIDENCIAL.
# Este archivo contiene vistas y endpoints que utilizan el motor de IA.
#
# Consulta el archivo LICENSE en la raíz del repositorio para más detalles
# sobre la protección de la Propiedad Intelectual de eGarage.
# -----------------------------------------------------------------------------

# ==================== EXPORTAR MECÁNICOS A EXCEL ====================

# import openpyxl
# from openpyxl.utils import get_column_letter
from collections import defaultdict
from datetime import date, datetime, timedelta
from decimal import Decimal

from django.db.models import (
    Count,
    DecimalField,
    ExpressionWrapper,
    F,
    FloatField,
    IntegerField,
    Q,
    Sum,
    Value,
)
from django.core.exceptions import ObjectDoesNotExist
from django.db.models.functions import Coalesce
from django.http import HttpResponse, JsonResponse, Http404
from django.shortcuts import render
from django.utils import timezone
from django.utils.dateparse import parse_date

from taller.auth.decorators import login_required_default
from taller.middleware.rate_limiting import rate_limit
from taller.models import ConfiguracionEmpresa, Documento
from taller.models.clientes import Cliente
from taller.models.lineas_documento import (
    LineaOtroServicio,
    LineaRepuesto,
    LineaServicio,
)
from taller.models.tecnico import Tecnico
from taller.models.vehiculos import Vehiculo
from taller.utils.empresa import get_or_create_empresa, get_user_empresa_safe
# LAZY IMPORT IA PARA EVITAR BLOQUEOS WINDOWS/PANDAS
from taller.reportes.kilometraje_reportes import ReporteKilometraje

# from taller.utils import get_or_create_empresa  # Eliminado: usamos la función local


# @login_required_default
# TEMPORALMENTE COMENTADO POR PROBLEMA DE MEMORIA CON OPENPYXL
# def exportar_mecanicos_excel(request):
#     """Exporta a Excel el reporte de mecánicos en el rango filtrado"""
#     empresa = get_user_empresa_safe(request.user)

#     # Obtener filtros (igual que en reportes_mecanicos)
#     fecha_desde = request.GET.get("fecha_desde") or (
#         date.today() - timedelta(days=30)
#     ).strftime("%Y-%m-%d")
#     fecha_hasta = request.GET.get("fecha_hasta") or date.today().strftime("%Y-%m-%d")
#     mecanico_id = request.GET.get("mecanico_id")

#     documentos_qs = Documento.objects.filter(
#         fecha_emision__range=[fecha_desde, fecha_hasta],
#         tecnico_responsable__isnull=False,
#         empresa=empresa,
#         tipo="FAC",  # usamos código de factura
#     ).select_related("tecnico_responsable")

#     if mecanico_id and mecanico_id != "todos":
#         documentos_qs = documentos_qs.filter(tecnico_responsable_id=mecanico_id)

#     # Crear libro Excel
#     wb = openpyxl.Workbook()
#     ws = wb.active
#     ws.title = "Reporte Mecánicos"

#     # Encabezados
#     headers = ["Mecánico", "Fecha", "Cliente", "Vehículo", "Monto Total"]
#     ws.append(headers)

#     # Ajustar ancho de columnas
#     for col_num, header in enumerate(headers, 1):
#         col_letter = get_column_letter(col_num)
#         ws.column_dimensions[col_letter].width = 20

# Rellenar datos
#     for doc in documentos_qs:
#         ws.append(
#             [
#                 doc.tecnico_responsable.nombre,
#                 doc.fecha_emision.strftime("%Y-%m-%d"),
#                 str(doc.cliente),
#                 str(doc.vehiculo),
#                 doc.total,
#             ]
#         )

# Respuesta HTTP
#     response = HttpResponse(
#         content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#     )
#     filename = f"reporte_mecanicos_{fecha_desde}_a_{fecha_hasta}.xlsx"
#     response["Content-Disposition"] = f'attachment; filename="{filename}"'

#     wb.save(response)
#     return response


# from collections import defaultdict
# from datetime import datetime
# from decimal import Decimal

# from django.db.models import FloatField
# from django.http import JsonResponse
# from django.shortcuts import render
# from django.utils import timezone

# from taller.models.clientes import Cliente
# from taller.models.documento import Documento
# from taller.models.lineas_documento import (
#     LineaOtroServicio,
#     LineaRepuesto,
#     LineaServicio,
# )
# from taller.models.tecnico import Tecnico
# from taller.models.vehiculos import Vehiculo
# from taller.utils.empresa import get_or_create_empresa
# # LAZY IMPORT IA PARA EVITAR BLOQUEOS WINDOWS/PANDAS


@login_required_default
def reportes_dashboard(request):
    return render(request, "taller/reportes/reportes.html")


@login_required_default
def reporte_repuestos(request):
    from taller.models.repuesto import Repuesto

    # 🔒 FILTRO CRÍTICO POR EMPRESA
    empresa = get_user_empresa_safe(request.user)

    # Top 10 repuestos más vendidos - FILTRADO POR EMPRESA
    repuesto_ventas = (
        LineaRepuesto.objects.filter(
            documento__tipo="FAC",
            documento__empresa=empresa,  # 🔒 FILTRO EMPRESA
        )
        .values("codigo", "nombre")
        .annotate(
            cantidad_total=Sum("cantidad"),
            ingresos=Sum(
                ExpressionWrapper(
                    F("cantidad") * F("precio_unitario"),
                    output_field=FloatField(),
                )
            ),
        )
        .order_by("-cantidad_total", "-documento__fecha_emision")
    )
    top_repuestos = list(repuesto_ventas[:10])

    # Repuestos con mayor margen de ganancia - FILTRADO POR EMPRESA
    # Márgenes solo si existen los campos en el modelo Repuesto

    repuestos_qs = Repuesto.objects.filter(empresa=empresa)
    total_precio_compra = Decimal("0.00")
    total_precio_venta = Decimal("0.00")
    ganancia_total = Decimal("0.00")

    if hasattr(Repuesto, "precio_compra") and hasattr(Repuesto, "precio_venta"):
        totales_globales = repuestos_qs.aggregate(
            total_precio_compra=Coalesce(
                Sum("precio_compra"),
                Value(Decimal("0.00")),
                output_field=DecimalField(max_digits=14, decimal_places=2),
            ),
            total_precio_venta=Coalesce(
                Sum("precio_venta"),
                Value(Decimal("0.00")),
                output_field=DecimalField(max_digits=14, decimal_places=2),
            ),
        )
        total_precio_compra = totales_globales["total_precio_compra"]
        total_precio_venta = totales_globales["total_precio_venta"]
        ganancia_total = total_precio_venta - total_precio_compra

        top_margen = list(
            repuestos_qs.annotate(
                total=Coalesce(
                    ExpressionWrapper(
                        Sum("linearepuesto__cantidad"),
                        output_field=FloatField(),
                    ),
                    Value(0.0, output_field=FloatField()),
                ),
                ingresos=Coalesce(
                    ExpressionWrapper(
                        Sum(
                            ExpressionWrapper(
                                F("linearepuesto__cantidad") * F("linearepuesto__precio_unitario"),
                                output_field=FloatField(),
                            )
                        ),
                        output_field=FloatField(),
                    ),
                    Value(0.0, output_field=FloatField()),
                ),
                margen=ExpressionWrapper(
                    (F("precio_venta") - F("precio_compra")) / F("precio_compra") * Value(100.0),
                    output_field=FloatField(),
                ),
            )
            .filter(total__gt=0, precio_compra__gt=0)
            .values("part_number", "nombre", "margen", "ingresos")
            .order_by("-margen")[:10]
        )
    else:
        top_margen = []

    # Repuestos con bajo stock - FILTRADO POR EMPRESA
    bajo_stock = Repuesto.objects.filter(
        cantidad_stock__lte=5,
        empresa=empresa,  # 🔒 FILTRO EMPRESA
    ).values("part_number", "nombre", "cantidad_stock")
    bajo_stock = [
        {
            "codigo": r["part_number"],
            "nombre": r["nombre"],
            "stock": r["cantidad_stock"],
        }
        for r in bajo_stock
    ]

    # Histórico de ventas mensuales - FILTRADO POR EMPRESA
    ventas = (
        LineaRepuesto.objects.filter(
            documento__tipo="FAC",
            documento__empresa=empresa,  # 🔒 FILTRO EMPRESA
        )
        .annotate(mes=F("documento__fecha_emision"))
        .values("mes")
        .annotate(total=Sum("cantidad"))
        .order_by("mes")
    )
    ventas_mensuales = defaultdict(int)
    for v in ventas:
        mes = v["mes"].strftime("%Y-%m") if v["mes"] else "Sin fecha"
        ventas_mensuales[mes] += v["total"]
    labels = list(ventas_mensuales.keys())
    data = list(ventas_mensuales.values())

    # Repuestos nunca vendidos - FILTRADO POR EMPRESA
    vendidos_codigos = set(
        LineaRepuesto.objects.filter(
            documento__tipo="FAC",
            documento__empresa=empresa,  # 🔒 FILTRO EMPRESA
        ).values_list("codigo", flat=True)
    )
    nunca_vendidos = Repuesto.objects.filter(empresa=empresa).exclude(
        part_number__in=vendidos_codigos
    )  # 🔒 FILTRO EMPRESA
    nunca_vendidos = [{"codigo": r.part_number, "nombre": r.nombre} for r in nunca_vendidos]

    context = {
        "top_repuestos": top_repuestos,
        "top_margen": top_margen,
        "bajo_stock": bajo_stock,
        "ventas_mensuales": {"labels": labels, "data": data},
        "nunca_vendidos": nunca_vendidos,
        "total_precio_compra": total_precio_compra,
        "total_precio_venta": total_precio_venta,
        "ganancia_total": ganancia_total,
    }
    return render(request, "taller/reportes/reporte_repuestos.html", context)


@login_required_default
def reporte_servicios(request):
    # 🔒 FILTRO CRÍTICO POR EMPRESA
    empresa = get_user_empresa_safe(request.user)

    tipos_facturacion = ["FAC", "BOL"]

    today = date.today()
    fecha_desde_str = request.GET.get("fecha_desde")
    fecha_hasta_str = request.GET.get("fecha_hasta")
    fecha_desde = parse_date(fecha_desde_str) if fecha_desde_str else today - timedelta(days=30)
    fecha_hasta = parse_date(fecha_hasta_str) if fecha_hasta_str else today
    if fecha_desde > fecha_hasta:
        fecha_desde, fecha_hasta = fecha_hasta, fecha_desde

    # Panel de facturación - FILTRADO POR EMPRESA
    lineas_base = LineaServicio.objects.filter(
        documento__empresa=empresa, documento__tipo__in=tipos_facturacion
    )

    # Total
    facturacion_total = (
        lineas_base.aggregate(
            total=Sum(
                ExpressionWrapper(F("precio_unitario") * F("cantidad"), output_field=DecimalField())
            )
        )["total"]
        or 0
    )

    # Por periodo (últimos 6 meses) - FILTRADO POR EMPRESA
    hoy = date.today()
    hace_6_meses = hoy - timedelta(days=180)
    facturacion_periodo_qs = (
        lineas_base.filter(documento__fecha_emision__gte=hace_6_meses)
        .annotate(
            total_calculado=ExpressionWrapper(
                F("precio_unitario") * F("cantidad"), output_field=DecimalField()
            ),
            mes=F("documento__fecha_emision"),
        )
        .values("mes")
        .annotate(total=Sum("total_calculado"))
        .order_by("mes")
    )
    facturacion_periodo = [
        {
            "mes": f["mes"].strftime("%Y-%m") if f["mes"] else "Sin fecha",
            "total": f["total"],
        }
        for f in facturacion_periodo_qs
    ]

    # Por servicio - FILTRADO POR EMPRESA
    facturacion_servicio = list(
        lineas_base.annotate(
            total_calculado=ExpressionWrapper(
                F("precio_unitario") * F("cantidad"), output_field=DecimalField()
            )
        )
        .values("nombre")
        .annotate(total=Sum("total_calculado"))
        .order_by("-total")[:10]
    )
    for f in facturacion_servicio:
        f["nombre"] = f["nombre"] or "Sin nombre"

    # Por cliente - FILTRADO POR EMPRESA
    facturacion_cliente_qs = (
        lineas_base.annotate(
            total_calculado=ExpressionWrapper(
                F("precio_unitario") * F("cantidad"), output_field=DecimalField()
            )
        )
        .values("documento__cliente__nombre", "documento__cliente__apellido")
        .annotate(total=Sum("total_calculado"))
        .order_by("-total")[:10]
    )
    facturacion_cliente = [
        {
            "cliente": (f["documento__cliente__nombre"] or "")
            + (
                " " + f["documento__cliente__apellido"] if f["documento__cliente__apellido"] else ""
            ),
            "total": f["total"],
        }
        for f in facturacion_cliente_qs
    ]

    Servicio = None
    try:
        from taller.servicios.models import Servicio
    except ImportError:
        pass

    # Top 10 servicios más vendidos - FILTRADO POR EMPRESA
    top_servicios = (
        lineas_base.annotate(
            total_calculado=ExpressionWrapper(
                F("precio_unitario") * F("cantidad"), output_field=DecimalField()
            )
        )
        .values("nombre")
        .annotate(cantidad=Count("id"), total=Sum("total_calculado"))
        .order_by("-cantidad")[:10]
    )

    # Servicios con mayor facturación - FILTRADO POR EMPRESA
    top_facturacion = (
        lineas_base.annotate(
            total_calculado=ExpressionWrapper(
                F("precio_unitario") * F("cantidad"), output_field=DecimalField()
            )
        )
        .values("nombre")
        .annotate(total=Sum("total_calculado"), cantidad=Count("id"))
        .order_by("-total")[:10]
    )

    # Histórico de servicios mensuales - FILTRADO POR EMPRESA
    servicios_mes = (
        lineas_base.annotate(
            total_calculado=ExpressionWrapper(
                F("precio_unitario") * F("cantidad"), output_field=DecimalField()
            ),
            mes=F("documento__fecha_emision"),
        )
        .values("mes")
        .annotate(total=Sum("total_calculado"))
        .order_by("mes")
    )
    historico = defaultdict(int)
    for s in servicios_mes:
        mes = s["mes"].strftime("%Y-%m") if s["mes"] else "Sin fecha"
        historico[mes] += s["total"]
    labels = list(historico.keys())
    data = list(historico.values())

    # Servicios nunca vendidos (si existe catálogo de servicios)
    nunca_vendidos = []
    if Servicio:
        vendidos_nombres = set(LineaServicio.objects.values_list("nombre", flat=True))
        nunca_vendidos = Servicio.objects.exclude(nombre__in=vendidos_nombres)
        nunca_vendidos = [{"nombre": s.nombre} for s in nunca_vendidos]

    # Rankings de vehículos atendidos

    # Solo documentos con servicios realizados
    doc_ids = lineas_base.values_list("documento_id", flat=True)
    vehiculos = Vehiculo.objects.filter(documentos__id__in=doc_ids).distinct()
    # Ranking de modelos
    ranking_modelos = (
        vehiculos.values("modelo__nombre").annotate(cantidad=Count("id")).order_by("-cantidad")[:10]
    )
    ranking_modelos = [
        {"modelo": r["modelo__nombre"] or "Sin modelo", "cantidad": r["cantidad"]}
        for r in ranking_modelos
    ]
    # Ranking de marcas
    ranking_marcas = (
        vehiculos.values("marca__nombre").annotate(cantidad=Count("id")).order_by("-cantidad")[:10]
    )
    ranking_marcas = [
        {"marca": r["marca__nombre"] or "Sin marca", "cantidad": r["cantidad"]}
        for r in ranking_marcas
    ]
    # Ranking de años
    ranking_anios = (
        vehiculos.values("anio").annotate(cantidad=Count("id")).order_by("-cantidad")[:10]
    )
    ranking_anios = [
        {"anio": r["anio"] or "Sin año", "cantidad": r["cantidad"]} for r in ranking_anios
    ]
    # Ranking de clientes frecuentes
    ranking_clientes = (
        vehiculos.values("cliente__nombre", "cliente__apellido")
        .annotate(cantidad=Count("id"))
        .order_by("-cantidad")[:10]
    )
    ranking_clientes = [
        {
            "cliente": (r["cliente__nombre"] or "")
            + (" " + r["cliente__apellido"] if r["cliente__apellido"] else ""),
            "cantidad": r["cantidad"],
        }
        for r in ranking_clientes
    ]

    # Panel de clientes
    hoy = date.today()
    hace_3_meses = hoy - timedelta(days=90)
    # Clientes activos: con más documentos en los últimos 6 meses
    clientes_activos = list(
        Cliente.objects.filter(
            documentos__fecha_emision__gte=hoy - timedelta(days=180),
            documentos__tipo__in=tipos_facturacion,
        )
        .annotate(cantidad=Count("documentos"))
        .order_by("-cantidad")
        .values("nombre", "apellido", "cantidad")[:10]
    )
    for c in clientes_activos:
        c["nombre"] = (c["nombre"] or "") + (" " + c["apellido"] if c["apellido"] else "")
    # Clientes nuevos: creados en los últimos 3 meses
    clientes_nuevos = list(
        Cliente.objects.filter(
            documentos__fecha_emision__gte=hace_3_meses, documentos__tipo__in=tipos_facturacion
        )
        .annotate(fecha_alta=F("documentos__fecha_emision"))
        .order_by("-fecha_alta")
        .values("nombre", "apellido", "fecha_alta")
        .distinct()[:10]
    )
    for c in clientes_nuevos:
        c["nombre"] = (c["nombre"] or "") + (" " + c["apellido"] if c["apellido"] else "")
    # Clientes históricos: con más documentos en total
    clientes_historicos = list(
        Cliente.objects.filter(documentos__tipo__in=tipos_facturacion)
        .annotate(cantidad=Count("documentos"))
        .order_by("-cantidad")
        .values("nombre", "apellido", "cantidad")[:10]
    )
    for c in clientes_historicos:
        c["nombre"] = (c["nombre"] or "") + (" " + c["apellido"] if c["apellido"] else "")
    # Clientes recurrentes: más de 1 documento en el último año
    clientes_recurrentes = list(
        Cliente.objects.filter(
            documentos__fecha_emision__gte=hoy - timedelta(days=365),
            documentos__tipo__in=tipos_facturacion,
        )
        .annotate(cantidad=Count("documentos"))
        .filter(cantidad__gt=1)
        .order_by("-cantidad")
        .values("nombre", "apellido", "cantidad")[:10]
    )
    for c in clientes_recurrentes:
        c["nombre"] = (c["nombre"] or "") + (" " + c["apellido"] if c["apellido"] else "")

    # Agenda y turnos: próximos turnos (documentos con fecha futura)
    turnos_proximos = list(
        Documento.objects.filter(fecha_emision__gte=hoy)
        .order_by("fecha_emision")
        .annotate(tipo_doc=F("tipo"))  # Cambiado para evitar conflicto con propiedad
        .values(
            "fecha_emision",
            "cliente__nombre",
            "cliente__apellido",
            "vehiculo__patente",
            "vehiculo__marca__nombre",
            "vehiculo__modelo__nombre",
            "tipo_doc",
        )[:20]
    )
    for t in turnos_proximos:
        t["cliente"] = (t["cliente__nombre"] or "") + (
            " " + t["cliente__apellido"] if t["cliente__apellido"] else ""
        )
        t["vehiculo"] = (
            (t["vehiculo__marca__nombre"] or "")
            + " "
            + (t["vehiculo__modelo__nombre"] or "")
            + " ("
            + (t["vehiculo__patente"] or "-")
            + ")"
        )

    lineas_filtradas = lineas_base.filter(
        documento__fecha_emision__range=[fecha_desde, fecha_hasta]
    )
    total_periodo_servicios = lineas_filtradas.aggregate(
        total=Sum(
            ExpressionWrapper(
                F("precio_unitario") * F("cantidad"),
                output_field=DecimalField(max_digits=14, decimal_places=2),
            )
        )
    )["total"] or Decimal("0.00")

    documentos_periodo = (
        Documento.objects.filter(
            empresa=empresa,
            tipo__in=tipos_facturacion,
            fecha_emision__range=[fecha_desde, fecha_hasta],
        )
        .annotate(
            total_servicios=Coalesce(
                Sum(
                    ExpressionWrapper(
                        F("lineas_servicio__cantidad") * F("lineas_servicio__precio_unitario"),
                        output_field=DecimalField(max_digits=14, decimal_places=2),
                    )
                ),
                Value(Decimal("0.00"), output_field=DecimalField(max_digits=14, decimal_places=2)),
            )
        )
        .filter(total_servicios__gt=0)
        .select_related("cliente", "tecnico_responsable")
        .order_by("-fecha_emision")
    )
    context = {
        "top_servicios": top_servicios,
        "top_facturacion": top_facturacion,
        "historico_servicios": {"labels": labels, "data": data},
        "nunca_vendidos": nunca_vendidos,
        "ranking_modelos": ranking_modelos,
        "ranking_marcas": ranking_marcas,
        "ranking_anios": ranking_anios,
        "ranking_clientes": ranking_clientes,
        "clientes_activos": clientes_activos,
        "clientes_nuevos": clientes_nuevos,
        "clientes_historicos": clientes_historicos,
        "clientes_recurrentes": clientes_recurrentes,
        "facturacion_total": facturacion_total,
        "facturacion_periodo": facturacion_periodo,
        "facturacion_servicio": facturacion_servicio,
        "facturacion_cliente": facturacion_cliente,
        "turnos_proximos": turnos_proximos,
        "fecha_desde": fecha_desde,
        "fecha_hasta": fecha_hasta,
        "total_periodo_servicios": total_periodo_servicios,
        "documentos_periodo": documentos_periodo,
    }
    return render(request, "taller/reportes/reporte_servicios.html", context)


@login_required_default
def dashboard_inteligencia_operativa(request):
    """
    🚀 Centro de Inteligencia Operativa - Dashboard Futurista 360°
    Análisis predictivo y KPIs avanzados para talleres automotrices
    """
    from decimal import InvalidOperation

    # 🔒 FILTRO CRÍTICO POR EMPRESA - Usuario autenticado garantizado
    empresa = get_user_empresa_safe(request.user)

    # 📊 Calcular KPIs principales - FILTRADO POR EMPRESA
    facturacion_total = (
        LineaServicio.objects.filter(
            documento__tipo="FAC",  # o el valor real de factura en tu choices
            documento__empresa=empresa,  # 🔒 FILTRO EMPRESA
        ).aggregate(
            total=Sum(
                ExpressionWrapper(F("precio_unitario") * F("cantidad"), output_field=DecimalField())
            )
        )[
            "total"
        ]
        or 0
    )

    # Calcular métricas adicionales para KPIs - FILTRADO POR EMPRESA
    total_documentos = Documento.objects.filter(
        tipo="FAC",
        empresa=empresa,  # 🔒 FILTRO EMPRESA
    ).count()
    total_clientes = Cliente.objects.filter(empresa=empresa).count()  # 🔒 FILTRO EMPRESA
    total_vehiculos = Vehiculo.objects.filter(empresa=empresa).count()  # 🔒 FILTRO EMPRESA

    # Facturación del mes actual vs mes anterior - FILTRADO POR EMPRESA
    hoy = timezone.now().date()
    inicio_mes_actual = hoy.replace(day=1)
    if inicio_mes_actual.month == 1:
        inicio_mes_anterior = inicio_mes_actual.replace(year=inicio_mes_actual.year - 1, month=12)
    else:
        inicio_mes_anterior = inicio_mes_actual.replace(month=inicio_mes_actual.month - 1)

    facturacion_mes_actual = (
        LineaServicio.objects.filter(
            documento__tipo="FAC",
            documento__fecha_emision__gte=inicio_mes_actual,
            documento__empresa=empresa,  # 🔒 FILTRO EMPRESA
        ).aggregate(total=Sum("precio_unitario"))["total"]
        or 0
    )

    facturacion_mes_anterior = (
        LineaServicio.objects.filter(
            documento__tipo="FAC",
            documento__fecha_emision__gte=inicio_mes_anterior,
            documento__fecha_emision__lt=inicio_mes_actual,
            documento__empresa=empresa,  # 🔒 FILTRO EMPRESA
        ).aggregate(total=Sum("precio_unitario"))["total"]
        or 0
    )

    # Calcular datos para gráficos de facturación mensual - FILTRADO POR EMPRESA
    facturacion_por_mes = []
    meses = [
        "Ene",
        "Feb",
        "Mar",
        "Abr",
        "May",
        "Jun",
        "Jul",
        "Ago",
        "Sep",
        "Oct",
        "Nov",
        "Dic",
    ]

    for i in range(6):  # Últimos 6 meses
        fecha_mes = hoy.replace(day=1) - timedelta(days=30 * (5 - i))
        inicio_mes = fecha_mes.replace(day=1)
        if inicio_mes.month == 12:
            fin_mes = inicio_mes.replace(year=inicio_mes.year + 1, month=1)
        else:
            fin_mes = inicio_mes.replace(month=inicio_mes.month + 1)

        facturacion = (
            LineaServicio.objects.filter(
                documento__tipo="FAC",
                documento__fecha_emision__gte=inicio_mes,
                documento__fecha_emision__lt=fin_mes,
                documento__empresa=empresa,  # 🔒 FILTRO EMPRESA
            ).aggregate(total=Sum("precio_unitario"))["total"]
            or 0
        )

        facturacion_por_mes.append(
            {"mes": meses[inicio_mes.month - 1], "valor": float(facturacion)}
        )

    # Datos para gráfico de servicios más demandados - FILTRADO POR EMPRESA
    servicios_demandados = (
        LineaServicio.objects.filter(
            documento__tipo="FAC",
            documento__empresa=empresa,  # 🔒 FILTRO EMPRESA
        )
        .values("nombre")
        .annotate(total_servicios=Count("id"), total_ingresos=Sum("precio_unitario"))
        .order_by("-total_servicios")[:5]
    )

    # Datos para mapa térmico (servicios por día de la semana) - FILTRADO POR EMPRESA
    servicios_por_dia = []
    for i in range(28):  # Últimas 4 semanas
        fecha = hoy - timedelta(days=i)
        servicios_dia = Documento.objects.filter(
            tipo="FAC",
            fecha_emision=fecha,
            empresa=empresa,  # 🔒 FILTRO EMPRESA
        ).count()
        servicios_por_dia.append(
            {
                "fecha": fecha,
                "servicios": servicios_dia,
                "dia_semana": fecha.weekday(),  # 0=Lunes, 6=Domingo
            }
        )

    # Clientes que no han vuelto en 60 días - FILTRADO POR EMPRESA
    hace_60_dias = hoy - timedelta(days=60)
    tipos_facturacion_local = ["FAC", "BOL"]
    clientes_inactivos = (
        Cliente.objects.filter(
            documentos__fecha_emision__lt=hace_60_dias,
            documentos__tipo__in=tipos_facturacion_local,
            empresa=empresa,  # 🔒 FILTRO EMPRESA
        )
        .distinct()
        .count()
    )

    # 📈 Cálculo de variaciones para el template (sin filtros custom)
    def _to_decimal(x):
        try:
            return Decimal(str(x or 0))
        except (InvalidOperation, TypeError):
            return Decimal("0")

    act = _to_decimal(facturacion_mes_actual)
    ant = _to_decimal(facturacion_mes_anterior)

    # bandera para el branch del template
    crece = act > ant

    # % crecimiento cuando act > ant  => (act/ant)*100 - 100
    if ant > 0:
        pct_up = float((act / ant) * Decimal("100") - Decimal("100"))
    else:
        # si no hay mes anterior (0) y hay ventas actuales, se considera 100%
        pct_up = 100.0 if act > 0 else 0.0

    # % disminución cuando ant > act  => (ant/act)*100 - 100
    if act > 0:
        pct_down = float((ant / act) * Decimal("100") - Decimal("100"))
    else:
        pct_down = 100.0 if ant > 0 else 0.0

    # 🚨 WIDGET: Recordatorios de Mantenimiento Urgentes
    recordatorios_urgentes = []
    total_recordatorios_urgentes = 0
    try:
        reporte_km = ReporteKilometraje(empresa)
        recordatorios = reporte_km.recordatorios_mantenimiento(
            servicio_km=10000,  # Cambio de aceite cada 10k km
            margen_alerta=1000,  # Alertar cuando falten 1k km
        )
        recordatorios_urgentes = [r for r in recordatorios if r["urgencia"] == "alta"][:5]  # Top 5
        total_recordatorios_urgentes = len([r for r in recordatorios if r["urgencia"] == "alta"])
    except Exception:
        # Si hay error, continuar sin mostrar recordatorios
        pass

    # 🛡️ WIDGET: Garantías Potenciales Abiertas
    garantias_potenciales = 0
    try:
        # Buscar documentos recientes (últimos 90 días) que podrían tener garantías activas
        fecha_limite = timezone.now().date() - timedelta(days=90)

        documentos_recientes = (
            Documento.objects.filter(
                empresa=empresa,
                tipo__in=["OT", "PRES"],
                estado="EMITIDO",
                fecha_emision__gte=fecha_limite,
                vehiculo__isnull=False,
            )
            .select_related("vehiculo", "registro_kilometraje")
            .prefetch_related("vehiculo__historial_kilometraje")
        )

        # Contar documentos que tienen registro de kilometraje y podrían tener garantía activa
        for doc in documentos_recientes:
            try:
                registro_km = doc.registro_kilometraje
                if registro_km:
                    # Verificar si está dentro del límite de garantía (5,000 km)
                    km_registro = registro_km.kilometraje
                    km_actual = doc.vehiculo.kilometraje_actual if doc.vehiculo else 0
                    if km_actual > 0:
                        km_recorridos = km_actual - km_registro
                        if km_recorridos < 5000:  # Dentro del límite de garantía
                            garantias_potenciales += 1
            except (ObjectDoesNotExist, AttributeError):
                # Documento no tiene registro_kilometraje asociado
                continue
    except Exception:
        # Si hay error, continuar sin contar garantías
        pass

    context = {
        "facturacion_total": facturacion_total,
        "total_documentos": total_documentos,
        "total_clientes": total_clientes,
        "total_vehiculos": total_vehiculos,
        "facturacion_mes_actual": facturacion_mes_actual,
        "facturacion_mes_anterior": facturacion_mes_anterior,
        "crece": crece,
        "pct_up": pct_up,  # usar cuando crece
        "pct_down": pct_down,  # usar cuando disminuye
        "facturacion_por_mes": facturacion_por_mes,
        "servicios_demandados": list(servicios_demandados),
        "servicios_por_dia": servicios_por_dia,
        "clientes_inactivos": clientes_inactivos,
        # KPIs calculados
        "ticket_promedio": facturacion_total / max(total_documentos, 1),
        "clientes_activos_porcentaje": 68,  # Simulado por ahora
        "margen_promedio": 45,  # Simulado por ahora
        "ingresos_por_hora": 22500,  # Simulado por ahora
        "vehiculos_por_semana": 23,  # Simulado por ahora
        "satisfaccion_cliente": 4.8,  # Simulado por ahora
        # 🚨 Widgets de Kilometraje
        "recordatorios_urgentes": recordatorios_urgentes,
        "total_recordatorios_urgentes": total_recordatorios_urgentes,
        "garantias_potenciales": garantias_potenciales,
    }

    return render(request, "taller/reportes/dashboard_inteligencia_operativa.html", context)


@login_required_default
@rate_limit(action="ia_prediccion", per_ip=True, per_user=True)
def diagnostico_ia(request):
    """
    🧠 Diagnóstico por IA - Análisis Predictivo Avanzado
    Motor de inteligencia artificial para optimización de talleres automotrices
    Protegido con rate limiting estricto
    """
    # 🔒 FILTRO CRÍTICO POR EMPRESA (siempre autenticado y robusto)
    empresa = get_user_empresa_safe(request.user)

    # Obtener documentos para análisis - FILTRADO POR EMPRESA
    documentos = (
        Documento.objects.select_related("cliente", "vehiculo")
        .prefetch_related("lineas_servicio")
        .filter(empresa=empresa)  # 🔒 FILTRO EMPRESA
        .all()
    )

    # Inicializar motor de IA
    motor_ia = MotorDiagnosticoIA()

    # Realizar análisis completo
    resultados = motor_ia.analizar_servicios_completo(documentos)

    # Preparar contexto para template
    contexto = {
        "servicios_crecimiento": resultados["servicios_crecimiento"],
        "servicios_declive": resultados["servicios_declive"],
        "estacionalidad": resultados["estacionalidad"],
        "comparativa_mercado": resultados["comparativa_mercado"],
        "recomendaciones_ia": resultados["recomendaciones_ia"],
        "predicciones_ingresos": resultados["predicciones_ingresos"],
        "alertas_criticas": resultados["alertas_criticas"],
        "insights_ai": resultados["insights_ai"],
        "total_documentos": documentos.count(),
        "fecha_analisis": date.today().strftime("%d/%m/%Y"),
    }

    return render(request, "taller/reportes/diagnostico_ia.html", contexto)


# ==================== REPORTES POR MECÁNICO ====================


@login_required_default
def reportes_mecanicos(request):
    """Vista principal para reportes por mecánico con IA"""

    # 🔒 FILTRO CRÍTICO POR EMPRESA
    empresa = get_user_empresa_safe(request.user)

    # Guardia para verificar el país de la empresa vs la ruta usada
    path_info = request.path_info
    if path_info.startswith("/us/") and empresa.pais not in ("US", "USA"):
        print(f"DEBUG - ⚠️ Empresa {empresa.pais} accediendo a ruta USA: {path_info}")
    elif path_info.startswith("/cl/") and empresa.pais not in ("CL", "CHILE"):
        print(f"DEBUG - ⚠️ Empresa {empresa.pais} accediendo a ruta Chile: {path_info}")

    print(f"DEBUG - 🌍 Empresa: {empresa.nombre_taller}, País: {empresa.pais}, Ruta: {path_info}")

    # Obtener filtros
    fecha_desde = request.GET.get("fecha_desde")
    fecha_hasta = request.GET.get("fecha_hasta")
    mecanico_id = request.GET.get("mecanico_id") or request.GET.get(
        "tecnico_id"
    )  # Aceptar ambos parámetros

    print(
        f"DEBUG - Filtros recibidos: fecha_desde={fecha_desde}, fecha_hasta={fecha_hasta}, mecanico_id={mecanico_id}"
    )

    # Verificar si el técnico existe antes de filtrar
    if mecanico_id and mecanico_id != "todos":
        try:
            # Validar que sea un número válido
            mecanico_id_int = int(mecanico_id)
            tecnico_verificar = Tecnico.objects.get(pk=mecanico_id_int, empresa=empresa)
            print(f"DEBUG - ✅ Técnico encontrado: {tecnico_verificar.nombre} (ID: {mecanico_id})")
        except (ValueError, Tecnico.DoesNotExist):
            print(
                f"DEBUG - ⚠️ Técnico con ID {mecanico_id} no válido o no existe en la empresa {empresa.nombre_taller}"
            )
            tecnicos_disponibles = Tecnico.objects.filter(empresa=empresa)
            print(
                f"DEBUG - 📋 Técnicos disponibles: {[f'{t.nombre} (ID: {t.pk})' for t in tecnicos_disponibles]}"
            )
            # No retornamos error, simplemente usamos None para mostrar todos
            mecanico_id = None

    # Valores por defecto
    if not fecha_desde:
        fecha_desde = (date.today() - timedelta(days=30)).strftime("%Y-%m-%d")
    if not fecha_hasta:
        fecha_hasta = date.today().strftime("%Y-%m-%d")

    print(
        f"DEBUG - Filtros aplicados: fecha_desde={fecha_desde}, fecha_hasta={fecha_hasta}, mecanico_id={mecanico_id}"
    )

    dividir_por_tecnico = (
        ConfiguracionEmpresa.objects.filter(empresa=empresa)
        .values_list("dividir_por_tecnico", flat=True)
        .first()
        or False
    )

    # Base SaaS: todos los documentos validos del periodo. El filtro por tecnico se
    # aplica por documento o por linea segun dividir_por_tecnico, no descartando
    # documentos sin tecnico a nivel cabecera.
    documentos_base = Documento.objects.filter(
        fecha_emision__range=[fecha_desde, fecha_hasta],
        empresa=empresa,  # 🔒 FILTRO EMPRESA
    ).exclude(estado="ANULADO")

    line_total_expr = ExpressionWrapper(
        F("cantidad") * F("precio_unitario") * (1 - F("descuento") / 100),
        output_field=DecimalField(max_digits=14, decimal_places=2),
    )
    otro_total_expr = ExpressionWrapper(
        F("cantidad") * F("precio_cliente"),
        output_field=DecimalField(max_digits=14, decimal_places=2),
    )

    def lineas_servicio_base():
        return LineaServicio.objects.filter(documento__in=documentos_base).annotate(
            tecnico_efectivo_id=Coalesce(
                F("tecnico_responsable_id"),
                F("documento__tecnico_responsable_id"),
                output_field=IntegerField(),
            )
        )

    def lineas_repuesto_base():
        return LineaRepuesto.objects.filter(documento__in=documentos_base).annotate(
            tecnico_efectivo_id=Coalesce(
                F("tecnico_responsable_id"),
                F("documento__tecnico_responsable_id"),
                output_field=IntegerField(),
            )
        )

    def lineas_otro_base():
        return LineaOtroServicio.objects.filter(documento__in=documentos_base).annotate(
            tecnico_efectivo_id=Coalesce(
                F("tecnico_responsable_id"),
                F("documento__tecnico_responsable_id"),
                output_field=IntegerField(),
            )
        )

    def documentos_para_tecnico(tecnico_id):
        if dividir_por_tecnico:
            servicio_doc_ids = lineas_servicio_base().filter(
                tecnico_efectivo_id=tecnico_id
            ).values("documento_id")
            repuesto_doc_ids = lineas_repuesto_base().filter(
                tecnico_efectivo_id=tecnico_id
            ).values("documento_id")
            otro_doc_ids = lineas_otro_base().filter(
                tecnico_efectivo_id=tecnico_id
            ).values("documento_id")
            return documentos_base.filter(
                Q(tecnico_responsable_id=tecnico_id)
                | Q(id__in=servicio_doc_ids)
                | Q(id__in=repuesto_doc_ids)
                | Q(id__in=otro_doc_ids)
            ).distinct()

        return documentos_base.filter(tecnico_responsable_id=tecnico_id)

    documentos_qs = documentos_base.order_by("-fecha_emision", "-id")

    print(f"DEBUG - Documentos antes del filtro de técnico: {documentos_qs.count()}")

    if mecanico_id and mecanico_id != "todos":
        documentos_qs = documentos_para_tecnico(mecanico_id).order_by("-fecha_emision", "-id")
        print(
            f"DEBUG - Documentos después del filtro de técnico {mecanico_id}: {documentos_qs.count()}"
        )

    print(f"DEBUG - Query SQL: {documentos_qs.query}")

    # Métricas generales - YA FILTRADO POR EMPRESA
    total_documentos = documentos_qs.count()

    # Calcular total generado correctamente (incluyendo cantidad y descuentos)

    # Total de servicios
    servicios_qs = lineas_servicio_base()
    repuestos_qs = lineas_repuesto_base()
    otros_qs = lineas_otro_base()
    if mecanico_id and mecanico_id != "todos":
        if dividir_por_tecnico:
            servicios_qs = servicios_qs.filter(tecnico_efectivo_id=mecanico_id)
            repuestos_qs = repuestos_qs.filter(tecnico_efectivo_id=mecanico_id)
            otros_qs = otros_qs.filter(tecnico_efectivo_id=mecanico_id)
        else:
            servicios_qs = servicios_qs.filter(documento__tecnico_responsable_id=mecanico_id)
            repuestos_qs = repuestos_qs.filter(documento__tecnico_responsable_id=mecanico_id)
            otros_qs = otros_qs.filter(documento__tecnico_responsable_id=mecanico_id)

    total_servicios = servicios_qs.aggregate(total=Sum(line_total_expr))["total"] or 0

    # Total de repuestos
    total_repuestos = repuestos_qs.aggregate(total=Sum(line_total_expr))["total"] or 0

    # Total de otros servicios
    total_otros = otros_qs.aggregate(total=Sum(otro_total_expr))["total"] or 0

    # Calcular IVA (19% solo sobre repuestos según la lógica de negocio)
    iva_repuestos = total_repuestos * Decimal("0.19")

    # Total general incluyendo IVA
    total_generado = total_servicios + total_repuestos + total_otros + iva_repuestos

    promedio_por_documento = round(
        total_generado / total_documentos if total_documentos > 0 else 0, 0
    )

    # Obtener lista de técnicos para el filtro
    tecnicos = Tecnico.objects.filter(empresa=empresa).order_by("nombre")

    print(f"DEBUG - Técnicos encontrados: {tecnicos.count()}")
    for tecnico in tecnicos:
        print(f"DEBUG - Técnico: {tecnico.nombre} (ID: {tecnico.pk})")

    print(f"DEBUG - Contexto todos_mecanicos: {[t.nombre for t in tecnicos]}")

    # Generar datos detallados por mecánico
    mecanicos_data = []
    for tecnico in tecnicos:
        # Documentos del técnico en el período
        docs_tecnico = documentos_para_tecnico(tecnico.id)
        total_docs_tecnico = docs_tecnico.count()

        if total_docs_tecnico == 0:
            continue  # Saltar técnicos sin documentos

        # Servicios del técnico
        if dividir_por_tecnico:
            servicios_tecnico = lineas_servicio_base().filter(tecnico_efectivo_id=tecnico.id)
            repuestos_tecnico = lineas_repuesto_base().filter(tecnico_efectivo_id=tecnico.id)
            otros_tecnico = lineas_otro_base().filter(tecnico_efectivo_id=tecnico.id)
        else:
            servicios_tecnico = LineaServicio.objects.filter(documento__in=docs_tecnico)
            repuestos_tecnico = LineaRepuesto.objects.filter(documento__in=docs_tecnico)
            otros_tecnico = LineaOtroServicio.objects.filter(documento__in=docs_tecnico)

        total_servicios_count = servicios_tecnico.count()

        # Total generado por el técnico (incluyendo cantidad y descuentos)
        # Servicios del técnico
        total_servicios_tecnico = servicios_tecnico.aggregate(total=Sum(line_total_expr))[
            "total"
        ] or 0

        # Repuestos del técnico
        total_repuestos_tecnico = repuestos_tecnico.aggregate(total=Sum(line_total_expr))[
            "total"
        ] or 0

        # Otros servicios del técnico
        total_otros_tecnico = otros_tecnico.aggregate(total=Sum(otro_total_expr))["total"] or 0

        # Calcular IVA del técnico (19% solo sobre repuestos)
        iva_repuestos_tecnico = total_repuestos_tecnico * Decimal("0.19")

        # Total generado por el técnico incluyendo IVA
        total_generado_tecnico = (
            total_servicios_tecnico
            + total_repuestos_tecnico
            + total_otros_tecnico
            + iva_repuestos_tecnico
        )

        # Promedio por documento del técnico
        promedio_tecnico = round(
            (total_generado_tecnico / total_docs_tecnico if total_docs_tecnico > 0 else 0),
            0,
        )

        # Top servicios del técnico (simplificado para evitar errores de agregación)
        servicios_top = (
            servicios_tecnico.values("nombre")
            .annotate(cantidad=Count("id"))
            .order_by("-cantidad")[:5]
        )

        mecanicos_data.append(
            {
                "mecanico": tecnico,
                "total_documentos": total_docs_tecnico,
                "total_servicios": total_servicios_count,
                "total_generado": total_generado_tecnico,
                "promedio_por_documento": promedio_tecnico,
                "servicios_top": servicios_top,
            }
        )

    # Preparar servicios detallados si se selecciona un técnico específico
    servicios_detallados = []
    tecnico_nombre = None
    total_servicios_tecnico = 0
    promedio_diario_tecnico = 0
    eficiencia_tecnico = 95

    if mecanico_id and mecanico_id != "todos":
        try:
            tecnico = Tecnico.objects.get(id=mecanico_id, empresa=empresa)
            tecnico_nombre = tecnico.nombre

            # Obtener documentos del técnico
            docs_tecnico = documentos_para_tecnico(tecnico.id)

            # Obtener todos los servicios del técnico con información del documento
            servicios_detallados_qs = LineaServicio.objects.filter(documento__in=docs_tecnico)
            if dividir_por_tecnico:
                servicios_detallados_qs = lineas_servicio_base().filter(
                    tecnico_efectivo_id=tecnico.id
                )

            servicios_detallados = servicios_detallados_qs.select_related(
                "documento",
                "documento__cliente",
                "documento__vehiculo",
                "documento__vehiculo__marca",
                "documento__vehiculo__modelo",
            ).annotate(
                precio_total=line_total_expr
            )

            # Calcular total de servicios del técnico
            total_servicios_tecnico = (
                servicios_detallados.aggregate(total=Sum("precio_total"))["total"] or 0
            )

            # Calcular promedio diario
            dias_periodo = (
                datetime.strptime(fecha_hasta, "%Y-%m-%d")
                - datetime.strptime(fecha_desde, "%Y-%m-%d")
            ).days + 1
            promedio_diario_tecnico = (
                total_servicios_tecnico / dias_periodo if dias_periodo > 0 else 0
            )

        except Tecnico.DoesNotExist:
            pass

    # Preparar contexto
    context = {
        "documentos": documentos_qs,
        "total_documentos": total_documentos,
        "total_generado": total_generado,
        "promedio_por_documento": promedio_por_documento,
        "tecnicos": tecnicos,
        "todos_mecanicos": tecnicos,  # Alias para el template
        "mecanicos_data": mecanicos_data,  # Datos detallados por mecánico
        "fecha_desde": fecha_desde,
        "fecha_hasta": fecha_hasta,
        "mecanico_id": mecanico_id,
        "tecnico_seleccionado": mecanico_id or "todos",
        "mecanico_seleccionado": mecanico_id or "todos",
        # Datos para vista detallada de técnico
        "servicios_detallados": servicios_detallados,
        "tecnico_nombre": tecnico_nombre,
        "total_servicios_tecnico": total_servicios_tecnico,
        "promedio_diario_tecnico": promedio_diario_tecnico,
        "eficiencia_tecnico": eficiencia_tecnico,
    }

    return render(request, "taller/reportes/reportes_mecanicos.html", context)


# ==================== FUNCIONES ADICIONALES DE REPORTES MECÁNICOS ====================


@login_required_default
def generar_pdf_mecanico(request, mecanico_id):
    """Genera un PDF con el reporte del mecánico específico"""
    # Implementación básica para evitar errores de importación
    return HttpResponse("PDF en desarrollo", content_type="text/plain")


@login_required_default
def generar_resumen_whatsapp_mecanico(request, mecanico_id):
    """Genera un resumen para WhatsApp del mecánico específico"""

    return JsonResponse({"mensaje": "Función en desarrollo"})


@login_required_default
def api_mecanicos_chart_data(request):
    """API para datos de gráficos de mecánicos"""

    return JsonResponse({"data": [], "labels": []})


@login_required_default
def api_repuestos_chart_data(request):
    """API para datos de gráficos de repuestos"""

    return JsonResponse({"data": [], "labels": []})


@login_required_default
def reportes_por_fecha(request):
    """Reportes filtrados por fecha"""
    return render(request, "taller/reportes/reportes_por_fecha.html", {})


@login_required_default
def reportes_repuestos_fecha(request, desde, hasta):
    """Reportes de repuestos en un rango de fechas"""

    return JsonResponse({"desde": desde, "hasta": hasta, "data": []})


@login_required_default
def reportes_servicios_fecha(request, desde, hasta):
    """Reportes de servicios en un rango de fechas"""

    return JsonResponse({"desde": desde, "hasta": hasta, "data": []})


@login_required_default
def reportes_otros_servicios_fecha(request, desde, hasta):
    """Reportes de otros servicios en un rango de fechas"""

    return JsonResponse({"desde": desde, "hasta": hasta, "data": []})


@login_required_default
def centro_contable_chile(request):
    """
    🧮 Centro Contable Chile - Contador Virtual

    Vista especial para cierre contable con:
    - Resumen contable mensual
    - Alertas inteligentes de calidad de datos
    - Validación de consistencia contable
    - Flujo optimizado para el dueño del taller (3 clics)
    """
    from datetime import datetime, timedelta
    from decimal import Decimal

    empresa = get_user_empresa_safe(request.user)

    # Manejo de filtros de fecha avanzados
    hoy = date.today()

    # Botones rápidos
    periodo_rapido = request.GET.get("periodo_rapido", "")
    fecha_desde_str = request.GET.get("fecha_desde", "")
    fecha_hasta_str = request.GET.get("fecha_hasta", "")

    if periodo_rapido == "mes_actual":
        fecha_inicio = date(hoy.year, hoy.month, 1)
        if hoy.month == 12:
            fecha_fin = date(hoy.year + 1, 1, 1) - timedelta(days=1)
        else:
            fecha_fin = date(hoy.year, hoy.month + 1, 1) - timedelta(days=1)
    elif periodo_rapido == "mes_anterior":
        if hoy.month == 1:
            fecha_inicio = date(hoy.year - 1, 12, 1)
            fecha_fin = date(hoy.year, 1, 1) - timedelta(days=1)
        else:
            fecha_inicio = date(hoy.year, hoy.month - 1, 1)
            fecha_fin = date(hoy.year, hoy.month, 1) - timedelta(days=1)
    elif periodo_rapido == "año_actual":
        fecha_inicio = date(hoy.year, 1, 1)
        fecha_fin = date(hoy.year, 12, 31)
    elif fecha_desde_str and fecha_hasta_str:
        # Filtros personalizados
        fecha_inicio = parse_date(fecha_desde_str) or date(hoy.year, hoy.month, 1)
        fecha_fin = parse_date(fecha_hasta_str) or hoy
        if fecha_inicio > fecha_fin:
            fecha_inicio, fecha_fin = fecha_fin, fecha_inicio
    else:
        # Por defecto: mes actual
        mes_seleccionado = request.GET.get("mes", hoy.strftime("%Y-%m"))
        try:
            año, mes = map(int, mes_seleccionado.split("-"))
            fecha_inicio = date(año, mes, 1)
            if mes == 12:
                fecha_fin = date(año + 1, 1, 1) - timedelta(days=1)
            else:
                fecha_fin = date(año, mes + 1, 1) - timedelta(days=1)
        except:
            fecha_inicio = date(hoy.year, hoy.month, 1)
            if hoy.month == 12:
                fecha_fin = date(hoy.year + 1, 1, 1) - timedelta(days=1)
            else:
                fecha_fin = date(hoy.year, hoy.month + 1, 1) - timedelta(days=1)

    # === DATOS CONTABLES DEL MES ===
    # Tipos de documentos facturables
    tipos_facturacion = ["FAC", "BOL"]

    documentos_mes = Documento.objects.filter(
        empresa=empresa,
        fecha_emision__range=[fecha_inicio, fecha_fin],
        tipo__in=tipos_facturacion,  # Solo facturas y boletas
        estado="EMITIDO",
    ).select_related("cliente", "vehiculo")

    # Totales contables
    total_facturado = documentos_mes.aggregate(total=Coalesce(Sum("total"), Value(Decimal("0"))))[
        "total"
    ] or Decimal("0")

    total_neto_repuestos = documentos_mes.aggregate(
        total=Coalesce(Sum("neto_repuestos"), Value(Decimal("0")))
    )["total"] or Decimal("0")

    total_neto_servicios = documentos_mes.aggregate(
        total=Coalesce(Sum("neto_servicios"), Value(Decimal("0")))
    )["total"] or Decimal("0")

    total_neto_otros = documentos_mes.aggregate(
        total=Coalesce(Sum("neto_otros_servicios"), Value(Decimal("0")))
    )["total"] or Decimal("0")

    total_iva = documentos_mes.aggregate(total=Coalesce(Sum("tax_amount"), Value(Decimal("0"))))[
        "total"
    ] or Decimal("0")

    # Calcular neto total (sin IVA)
    neto_total = total_neto_repuestos + total_neto_servicios + total_neto_otros

    # === CÁLCULOS CONTABLES ESPECÍFICOS ===
    # Neto afecto = repuestos (base imponible 19%)
    neto_afecto = total_neto_repuestos
    # Neto exento = servicios (sin IVA en Chile)
    neto_exento = total_neto_servicios + total_neto_otros
    # IVA = 19% sobre neto afecto
    iva_periodo = total_iva
    # Total ventas = neto afecto + exento + IVA
    total_ventas = neto_afecto + neto_exento + iva_periodo

    # === VENTAS DIARIAS PARA GRÁFICO ===
    ventas_diarias_list = list(
        documentos_mes.values("fecha_emision")
        .annotate(total_dia=Coalesce(Sum("total"), Value(Decimal("0"))))
        .order_by("fecha_emision")
    )

    # Convertir a formato JSON-friendly
    import json
    from django.core.serializers.json import DjangoJSONEncoder

    ventas_diarias_json = json.dumps(
        [
            {"fecha_emision": str(v["fecha_emision"]), "total_dia": float(v["total_dia"])}
            for v in ventas_diarias_list
        ],
        cls=DjangoJSONEncoder,
    )

    # === ALERTAS INTELIGENTES ORGANIZADAS POR SEVERIDAD ===
    alertas_criticas = []
    alertas_advertencias = []
    alertas_info = []

    # 1. Documentos en borrador
    docs_borrador_list = Documento.objects.filter(
        empresa=empresa,
        fecha_emision__range=[fecha_inicio, fecha_fin],
        tipo__in=tipos_facturacion,
        estado="BORRADOR",
    )
    if docs_borrador_list.exists():
        alertas_criticas.append(
            {
                "tipo": "critica",
                "icono": "🔴",
                "titulo": f"{docs_borrador_list.count()} documento{{ docs_borrador_list.count()|pluralize }} en estado BORRADOR dentro del período",
                "mensaje": "Documentos en borrador que deben ser emitidos o eliminados antes del cierre.",
                "documentos": list(docs_borrador_list.values_list("id", flat=True)),
            }
        )

    # 2. Documentos sin número
    docs_sin_numero = documentos_mes.filter(numero__in=["", None])
    if docs_sin_numero.exists():
        alertas_criticas.append(
            {
                "tipo": "critica",
                "icono": "🔴",
                "titulo": f"{docs_sin_numero.count()} documento{{ docs_sin_numero.count()|pluralize }} sin número",
                "mensaje": "Documentos sin número de factura/boleta. Requiere atención inmediata.",
                "documentos": list(docs_sin_numero.values_list("id", flat=True)),
            }
        )

    # 3. Documentos sin cliente o sin RUT
    docs_sin_cliente = documentos_mes.filter(cliente__isnull=True)
    docs_sin_rut_list = documentos_mes.filter(cliente__isnull=False).filter(
        Q(cliente__tax_id__isnull=True) | Q(cliente__tax_id="")
    )

    if docs_sin_cliente.exists():
        alertas_criticas.append(
            {
                "tipo": "critica",
                "icono": "🔴",
                "titulo": f"{docs_sin_cliente.count()} documento{{ docs_sin_cliente.count()|pluralize }} sin cliente",
                "mensaje": "Documentos sin cliente asociado. Datos incompletos.",
                "documentos": list(docs_sin_cliente.values_list("id", flat=True)),
            }
        )

    if docs_sin_rut_list.exists():
        alertas_criticas.append(
            {
                "tipo": "critica",
                "icono": "🔴",
                "titulo": f"{docs_sin_rut_list.count()} documento{{ docs_sin_rut_list.count()|pluralize }} con cliente sin RUT",
                "mensaje": "Documentos con clientes que no tienen RUT registrado.",
                "documentos": list(docs_sin_rut_list.values_list("id", flat=True)),
            }
        )

    # 4. Inconsistencia en totales (IVA) - ADVERTENCIA
    iva_esperado = total_neto_repuestos * Decimal("0.19")
    diferencia_iva = abs(total_iva - iva_esperado)
    if diferencia_iva > Decimal("100"):  # Tolerancia de $100 CLP
        docs_inconsistentes = documentos_mes.filter(
            Q(
                tax_amount__lt=ExpressionWrapper(
                    F("neto_repuestos") * Decimal("0.19") - Decimal("50"),
                    output_field=DecimalField(),
                )
            )
            | Q(
                tax_amount__gt=ExpressionWrapper(
                    F("neto_repuestos") * Decimal("0.19") + Decimal("50"),
                    output_field=DecimalField(),
                )
            )
        )
        alertas_advertencias.append(
            {
                "tipo": "advertencia",
                "icono": "🟡",
                "titulo": "1 documento con totales inconsistentes (revisar líneas)",
                "mensaje": f"IVA calculado: ${total_iva:,.0f} | IVA esperado: ${iva_esperado:,.0f} | Diferencia: ${diferencia_iva:,.0f}",
                "documentos": (
                    list(docs_inconsistentes.values_list("id", flat=True))
                    if docs_inconsistentes.exists()
                    else []
                ),
            }
        )

    # 5. Documentos con total en cero - ADVERTENCIA
    docs_cero_list = documentos_mes.filter(total=Decimal("0"))
    if docs_cero_list.exists():
        alertas_advertencias.append(
            {
                "tipo": "advertencia",
                "icono": "🟡",
                "titulo": f"{docs_cero_list.count()} documento{{ docs_cero_list.count()|pluralize }} con total $0",
                "mensaje": "Documentos con monto total en cero. Verificar líneas de detalle.",
                "documentos": list(docs_cero_list.values_list("id", flat=True)),
            }
        )

    # 6. Documentos con tipo no definido - ADVERTENCIA
    docs_tipo_indefinido = documentos_mes.filter(tipo__isnull=True) | documentos_mes.filter(tipo="")
    if docs_tipo_indefinido.exists():
        alertas_advertencias.append(
            {
                "tipo": "advertencia",
                "icono": "🟡",
                "titulo": f'{docs_tipo_indefinido.count()} documento{{ docs_tipo_indefinido.count()|pluralize }} con tipo de documento no definido (usando "Interno")',
                "mensaje": 'Documentos sin tipo definido. Se están usando como "Interno".',
                "documentos": list(docs_tipo_indefinido.values_list("id", flat=True)),
            }
        )

    # 7. Documentos con descuento > 50% - ADVERTENCIA
    from taller.models.lineas_documento import LineaRepuesto, LineaServicio

    lineas_descuento_alto = (
        LineaRepuesto.objects.filter(
            documento__empresa=empresa,
            documento__fecha_emision__range=[fecha_inicio, fecha_fin],
            documento__tipo__in=tipos_facturacion,
            descuento__gt=50,
        )
        .values("documento_id")
        .distinct()
    )

    if lineas_descuento_alto.exists():
        alertas_advertencias.append(
            {
                "tipo": "advertencia",
                "icono": "🟡",
                "titulo": f"{lineas_descuento_alto.count()} documento{{ lineas_descuento_alto.count()|pluralize }} con descuento > 50% en una línea (revisar)",
                "mensaje": "Documentos con descuentos muy altos. Verificar que sean correctos.",
                "documentos": list(lineas_descuento_alto.values_list("documento_id", flat=True)),
            }
        )

    # 8. Documentos anulados - INFO
    docs_anulados_list = Documento.objects.filter(
        empresa=empresa,
        fecha_emision__range=[fecha_inicio, fecha_fin],
        tipo__in=tipos_facturacion,
        estado="ANULADO",
    )
    if docs_anulados_list.exists():
        alertas_info.append(
            {
                "tipo": "info",
                "icono": "🟢",
                "titulo": f"{docs_anulados_list.count()} documento{{ docs_anulados_list.count()|pluralize }} anulado{{ docs_anulados_list.count()|pluralize }} en el período (no se consideran en totales)",
                "mensaje": "Documentos anulados en el período. No se incluyen en los totales contables.",
                "documentos": list(docs_anulados_list.values_list("id", flat=True)),
            }
        )

    # Combinar todas las alertas para compatibilidad
    alertas = alertas_criticas + alertas_advertencias + alertas_info

    # === CALIDAD DE DATOS ===
    calidad_datos = {"score": 100, "problemas": []}

    total_docs = documentos_mes.count()
    if total_docs > 0:
        # Porcentaje de documentos con número
        docs_sin_numero_count = documentos_mes.filter(numero__in=["", None]).count()
        pct_con_numero = ((total_docs - docs_sin_numero_count) / total_docs) * 100
        # Porcentaje de documentos con cliente
        docs_sin_cliente_count = documentos_mes.filter(cliente__isnull=True).count()
        pct_con_cliente = ((total_docs - docs_sin_cliente_count) / total_docs) * 100
        # Porcentaje de documentos con total válido
        docs_cero_count = documentos_mes.filter(total=Decimal("0")).count()
        pct_con_total = ((total_docs - docs_cero_count) / total_docs) * 100

        # Score de calidad (promedio ponderado)
        calidad_datos["score"] = int((pct_con_numero + pct_con_cliente + pct_con_total) / 3)

        if pct_con_numero < 100:
            calidad_datos["problemas"].append(f"{100 - pct_con_numero:.0f}% sin número")
        if pct_con_cliente < 100:
            calidad_datos["problemas"].append(f"{100 - pct_con_cliente:.0f}% sin cliente")
        if pct_con_total < 100:
            calidad_datos["problemas"].append(f"{100 - pct_con_total:.0f}% con total $0")
    else:
        calidad_datos["score"] = 0
        calidad_datos["problemas"].append("No hay documentos en el período")

    # === RESUMEN POR TIPO DE DOCUMENTO ===
    resumen_tipos = (
        documentos_mes.values("tipo")
        .annotate(
            cantidad=Count("id"),
            total=Coalesce(Sum("total"), Value(Decimal("0"))),
            neto=Coalesce(
                Sum("neto_repuestos") + Sum("neto_servicios") + Sum("neto_otros_servicios"),
                Value(Decimal("0")),
            ),
            iva=Coalesce(Sum("tax_amount"), Value(Decimal("0"))),
        )
        .order_by("tipo")
    )

    # === ESTADÍSTICAS ADICIONALES ===
    total_documentos = documentos_mes.count()
    clientes_unicos = documentos_mes.values("cliente").distinct().count()
    promedio_ticket = total_facturado / total_documentos if total_documentos > 0 else Decimal("0")

    # Documentos en borrador
    docs_borrador = Documento.objects.filter(
        empresa=empresa,
        fecha_emision__range=[fecha_inicio, fecha_fin],
        tipo__in=tipos_facturacion,
        estado="BORRADOR",
    ).count()

    # Documentos sin RUT de cliente (usando tax_id)
    docs_sin_rut = (
        documentos_mes.filter(cliente__isnull=False)
        .filter(Q(cliente__tax_id__isnull=True) | Q(cliente__tax_id=""))
        .count()
    )

    # Desglose por repuestos y servicios
    from taller.models.lineas_documento import LineaRepuesto, LineaServicio

    lineas_repuestos = LineaRepuesto.objects.filter(
        documento__empresa=empresa,
        documento__fecha_emision__range=[fecha_inicio, fecha_fin],
        documento__tipo__in=tipos_facturacion,
        documento__estado="EMITIDO",
    )

    lineas_servicios = LineaServicio.objects.filter(
        documento__empresa=empresa,
        documento__fecha_emision__range=[fecha_inicio, fecha_fin],
        documento__tipo__in=tipos_facturacion,
        documento__estado="EMITIDO",
    )

    # Repuestos vendidos (detallado)
    repuestos_vendidos = (
        lineas_repuestos.values("codigo", "nombre")
        .annotate(
            cantidad_total=Sum("cantidad"),
            neto_total=Sum(
                ExpressionWrapper(F("cantidad") * F("precio_unitario"), output_field=DecimalField())
            ),
        )
        .order_by("-neto_total")
    )

    # Servicios realizados (detallado)
    servicios_realizados = (
        lineas_servicios.values("nombre")
        .annotate(
            cantidad_total=Sum("cantidad"),
            monto_total=Sum(
                ExpressionWrapper(F("cantidad") * F("precio_unitario"), output_field=DecimalField())
            ),
        )
        .order_by("-monto_total")
    )

    # Top repuestos (para preview)
    top_repuestos = repuestos_vendidos[:10]

    # Top servicios (para preview)
    top_servicios = servicios_realizados[:10]

    # Calcular neto por documento para el template
    documentos_con_neto = []
    for doc in documentos_mes[:100]:
        neto_doc = doc.neto_repuestos + doc.neto_servicios + doc.neto_otros_servicios
        neto_exento_doc = doc.neto_servicios + doc.neto_otros_servicios
        documentos_con_neto.append({"doc": doc, "neto": neto_doc, "neto_exento": neto_exento_doc})

    # === CONTEXTO ===
    context = {
        "empresa": empresa,
        "fecha_inicio": fecha_inicio,
        "fecha_fin": fecha_fin,
        "fecha_desde_str": fecha_desde_str or fecha_inicio.strftime("%Y-%m-%d"),
        "fecha_hasta_str": fecha_hasta_str or fecha_fin.strftime("%Y-%m-%d"),
        # Cálculos contables específicos
        "neto_afecto": neto_afecto,
        "neto_exento": neto_exento,
        "iva_periodo": iva_periodo,
        "total_ventas": total_ventas,
        # Datos para gráfico
        "ventas_diarias": ventas_diarias_json,
        # Totales generales (compatibilidad)
        "total_facturado": total_facturado,
        "total_neto_repuestos": total_neto_repuestos,
        "total_neto_servicios": total_neto_servicios,
        "total_neto_otros": total_neto_otros,
        "neto_total": neto_total,
        "total_iva": total_iva,
        # Alertas organizadas
        "alertas": alertas,
        "alertas_criticas": alertas_criticas,
        "alertas_advertencias": alertas_advertencias,
        "alertas_info": alertas_info,
        "calidad_datos": calidad_datos,
        "resumen_tipos": resumen_tipos,
        "total_documentos": total_documentos,
        "clientes_unicos": clientes_unicos,
        "promedio_ticket": promedio_ticket,
        "documentos_mes": documentos_mes[:100],  # Primeros 100 para libro de ventas
        "documentos_con_neto": documentos_con_neto,
        "docs_borrador": docs_borrador,
        "docs_sin_rut": docs_sin_rut,
        # Repuestos y servicios detallados
        "repuestos_vendidos": list(repuestos_vendidos),
        "servicios_realizados": list(servicios_realizados),
        "top_repuestos": list(top_repuestos),
        "top_servicios": list(top_servicios),
    }

    return render(request, "taller/reportes/centro_contable_chile.html", context)


# ==================== REPORTES DE KILOMETRAJE ====================


@login_required_default
def recordatorios_mantenimiento(request):
    """
    🚨 Vista de Recordatorios de Mantenimiento Predictivo

    Muestra vehículos que están cerca de necesitar mantenimiento,
    permitiendo al taller contactar proactivamente a los clientes.
    """
    # 🔒 FILTRO CRÍTICO POR EMPRESA
    empresa = get_user_empresa_safe(request.user)

    # Obtener parámetros de configuración desde GET (con valores por defecto)
    servicio_km = int(request.GET.get("servicio_km", 10000))  # Kilometraje del servicio
    margen_alerta = int(request.GET.get("margen_alerta", 1000))  # Margen de alerta en km
    tipo_servicio = request.GET.get("tipo_servicio", "cambio_aceite")  # Tipo de servicio

    # Configuraciones predefinidas por tipo de servicio
    configuraciones_servicio = {
        "cambio_aceite": {
            "servicio_km": 10000,
            "margen_alerta": 1000,
            "nombre": "Cambio de Aceite",
            "descripcion": "Servicio recomendado cada 10,000 km",
        },
        "revision_menor": {
            "servicio_km": 15000,
            "margen_alerta": 2000,
            "nombre": "Revisión Menor",
            "descripcion": "Revisión menor cada 15,000 km",
        },
        "revision_mayor": {
            "servicio_km": 50000,
            "margen_alerta": 5000,
            "nombre": "Revisión Mayor",
            "descripcion": "Revisión mayor cada 50,000 km",
        },
        "personalizado": {
            "servicio_km": servicio_km,
            "margen_alerta": margen_alerta,
            "nombre": "Personalizado",
            "descripcion": f"Servicio cada {servicio_km:,} km",
        },
    }

    # Obtener configuración del servicio seleccionado
    config = configuraciones_servicio.get(tipo_servicio, configuraciones_servicio["cambio_aceite"])

    # Si es personalizado, usar los valores de GET
    if tipo_servicio == "personalizado":
        config["servicio_km"] = servicio_km
        config["margen_alerta"] = margen_alerta

    # Crear instancia del reporte
    reporte = ReporteKilometraje(empresa)

    # Obtener recordatorios
    recordatorios = reporte.recordatorios_mantenimiento(
        servicio_km=config["servicio_km"], margen_alerta=config["margen_alerta"]
    )

    # Separar por urgencia
    recordatorios_urgentes = [r for r in recordatorios if r["urgencia"] == "alta"]
    recordatorios_medios = [r for r in recordatorios if r["urgencia"] == "media"]

    # Estadísticas
    total_recordatorios = len(recordatorios)
    total_urgentes = len(recordatorios_urgentes)
    total_medios = len(recordatorios_medios)

    # Preparar contexto
    context = {
        "recordatorios": recordatorios,
        "recordatorios_urgentes": recordatorios_urgentes,
        "recordatorios_medios": recordatorios_medios,
        "total_recordatorios": total_recordatorios,
        "total_urgentes": total_urgentes,
        "total_medios": total_medios,
        "configuraciones_servicio": configuraciones_servicio,
        "tipo_servicio_actual": tipo_servicio,
        "config_actual": config,
        "empresa": empresa,
    }

    return render(request, "taller/reportes/recordatorios_mantenimiento.html", context)


@login_required_default
def historial_mantenimiento_vehiculo(request, vehiculo_id):
    """
    📋 Vista de Historial de Mantenimiento Detallado de un Vehículo

    Muestra el historial completo tipo "Libro de Mantenciones Digital"
    """
    # 🔒 FILTRO CRÍTICO POR EMPRESA
    empresa = get_user_empresa_safe(request.user)

    # Obtener vehículo (con filtro de empresa)
    try:
        vehiculo = Vehiculo.objects.get(pk=vehiculo_id, empresa=empresa)
    except Vehiculo.DoesNotExist:
        raise Http404("Vehículo no encontrado")

    # Crear instancia del reporte
    reporte = ReporteKilometraje(empresa)

    # Obtener historial
    historial_data = reporte.historial_mantenimiento_vehiculo(vehiculo)

    # Preparar contexto
    context = {
        "historial_data": historial_data,
        "vehiculo": vehiculo,
        "empresa": empresa,
    }

    return render(request, "taller/reportes/historial_vehiculo.html", context)


@login_required_default
def api_historial_vehiculo(request, vehiculo_id):
    """
    📡 API Endpoint: Historial de Mantenimiento de un Vehículo

    Retorna el historial completo en formato JSON estructurado.
    Útil para Portal del Cliente y exportaciones.
    """
    # 🔒 FILTRO CRÍTICO POR EMPRESA
    empresa = get_user_empresa_safe(request.user)

    try:
        vehiculo = Vehiculo.objects.get(pk=vehiculo_id, empresa=empresa)
    except Vehiculo.DoesNotExist:
        return JsonResponse({"error": "Vehículo no encontrado"}, status=404)

    # Crear instancia del reporte
    reporte = ReporteKilometraje(empresa)

    # Exportar historial en formato estructurado
    historial_export = reporte.exportar_historial_vehiculo(vehiculo, formato="dict")

    # Serializar datos para JSON (convertir Decimal, datetime, etc.)
    def serialize_value(value):
        """Convierte valores no serializables a tipos básicos"""
        if isinstance(value, Decimal):
            return float(value)
        elif hasattr(value, "isoformat"):  # datetime, date
            return value.isoformat()
        elif hasattr(value, "__dict__"):  # Objetos modelo
            return str(value)
        return value

    def serialize_dict(d):
        """Recursivamente serializa un diccionario"""
        if isinstance(d, dict):
            return {k: serialize_dict(v) for k, v in d.items()}
        elif isinstance(d, list):
            return [serialize_dict(item) for item in d]
        else:
            return serialize_value(d)

    historial_serializado = serialize_dict(historial_export)

    return JsonResponse(historial_serializado, safe=False)


@login_required_default
def exportar_historial_pdf(request, vehiculo_id):
    """
    📄 Exportar Historial de Mantenimiento a PDF

    Genera un PDF del historial completo del vehículo usando WeasyPrint.
    """
    # 🔒 FILTRO CRÍTICO POR EMPRESA
    empresa = get_user_empresa_safe(request.user)

    try:
        vehiculo = Vehiculo.objects.get(pk=vehiculo_id, empresa=empresa)
    except Vehiculo.DoesNotExist:
        raise Http404("Vehículo no encontrado")

    # Crear instancia del reporte
    reporte = ReporteKilometraje(empresa)
    historial_data = reporte.historial_mantenimiento_vehiculo(vehiculo)

    try:
        from weasyprint import HTML
        from weasyprint.text.fonts import FontConfiguration
        from django.template.loader import render_to_string
        from django.utils import timezone

        # Preparar contexto
        context = {
            "historial_data": historial_data,
            "vehiculo": vehiculo,
            "empresa": empresa,
            "fecha_generacion": timezone.now(),
        }

        # Renderizar HTML
        html_string = render_to_string(
            "taller/reportes/historial_vehiculo_pdf.html", context, request=request
        )

        # Configuración de fuentes
        font_config = FontConfiguration()

        # Base URL para recursos (logos, imágenes)
        base_url = None
        if request:
            try:
                base_url = request.build_absolute_uri("/")
            except Exception:
                base_url = None

        # Generar PDF
        html = HTML(string=html_string, base_url=base_url)
        pdf_bytes = html.write_pdf(font_config=font_config)

        # Nombre del archivo
        patente_slug = vehiculo.patente.replace(" ", "_").replace("-", "_")
        filename = f"Historial_Mantenimiento_{patente_slug}_{timezone.now().strftime('%Y%m%d')}.pdf"

        # Retornar PDF como respuesta
        response = HttpResponse(pdf_bytes, content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    except ImportError:
        # Si WeasyPrint no está instalado, retornar error
        return HttpResponse(
            "Error: WeasyPrint no está instalado. Por favor, instálelo con: pip install weasyprint",
            status=500,
            content_type="text/plain",
        )
    except Exception as e:
        # Log del error y retornar mensaje
        import logging

        logger = logging.getLogger(__name__)
        logger.error(f"Error generando PDF de historial: {e}", exc_info=True)
        return HttpResponse(
            f"Error al generar PDF: {str(e)}", status=500, content_type="text/plain"
        )


@login_required_default
def exportar_historial_excel(request, vehiculo_id):
    """
    📊 Exportar Historial de Mantenimiento a Excel

    Genera un archivo Excel del historial completo del vehículo usando openpyxl.
    """
    # 🔒 FILTRO CRÍTICO POR EMPRESA
    empresa = get_user_empresa_safe(request.user)

    try:
        vehiculo = Vehiculo.objects.get(pk=vehiculo_id, empresa=empresa)
    except Vehiculo.DoesNotExist:
        raise Http404("Vehículo no encontrado")

    # Crear instancia del reporte
    reporte = ReporteKilometraje(empresa)
    historial_data = reporte.historial_mantenimiento_vehiculo(vehiculo)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from django.utils import timezone
        from io import BytesIO

        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Historial de Mantenimiento"

        # Estilos
        header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        title_font = Font(bold=True, size=14, color="1e40af")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Título
        ws.merge_cells("A1:E1")
        ws["A1"] = f"Historial de Mantenimiento - {vehiculo.patente}"
        ws["A1"].font = title_font
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        # Información del vehículo
        row = 3
        ws[f"A{row}"] = "Información del Vehículo"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        row += 1

        info_data = [
            ["Patente", vehiculo.patente],
            ["Marca / Modelo", f"{vehiculo.get_marca_display()} {vehiculo.get_modelo_display()}"],
            ["Año", vehiculo.anio],
            ["Kilometraje Actual", f"{vehiculo.kilometraje_actual} km"],
            ["Cliente", vehiculo.cliente.nombre if vehiculo.cliente else "N/A"],
        ]

        if vehiculo.vin:
            info_data.append(["VIN", vehiculo.vin])

        for label, value in info_data:
            ws[f"A{row}"] = label
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"B{row}"] = value
            row += 1

        row += 1

        # Resumen
        ws[f"A{row}"] = "Resumen"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        row += 1

        resumen_data = [
            ["Total Servicios", historial_data["resumen"]["total_servicios"]],
            ["Total Invertido", f"${historial_data['resumen']['total_gastado']:,.0f}"],
        ]

        if historial_data["resumen"].get("km_promedio_entre_servicios"):
            resumen_data.append(
                [
                    "KM Promedio entre Servicios",
                    f"{historial_data['resumen']['km_promedio_entre_servicios']:,.0f} km",
                ]
            )

        if historial_data["resumen"].get("dias_promedio_entre_servicios"):
            resumen_data.append(
                [
                    "Días Promedio entre Servicios",
                    f"{historial_data['resumen']['dias_promedio_entre_servicios']:.0f} días",
                ]
            )

        for label, value in resumen_data:
            ws[f"A{row}"] = label
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"B{row}"] = value
            row += 1

        row += 2

        # Encabezados de tabla
        headers = [
            "Fecha",
            "Documento",
            "Tipo",
            "Trabajos Realizados",
            "Kilometraje",
            "Monto",
            "Técnico",
        ]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        row += 1

        # Datos del historial
        for entrada in historial_data["historial"]:
            ws.cell(row=row, column=1).value = (
                entrada["fecha"].strftime("%d/%m/%Y") if entrada["fecha"] else ""
            )
            ws.cell(row=row, column=2).value = entrada["numero_documento"]
            ws.cell(row=row, column=3).value = entrada["tipo"]
            ws.cell(row=row, column=4).value = entrada["trabajos_realizados"]
            ws.cell(row=row, column=5).value = (
                f"{entrada['kilometraje']:,.0f} km" if entrada.get("kilometraje") else "N/A"
            )
            ws.cell(row=row, column=6).value = (
                f"${entrada['monto']:,.0f}" if entrada.get("monto") else "$0"
            )
            ws.cell(row=row, column=7).value = entrada.get("tecnico", "N/A")

            # Aplicar bordes
            for col in range(1, 8):
                ws.cell(row=row, column=col).border = border

            row += 1

        # Ajustar ancho de columnas
        column_widths = {
            "A": 12,  # Fecha
            "B": 15,  # Documento
            "C": 10,  # Tipo
            "D": 40,  # Trabajos
            "E": 15,  # Kilometraje
            "F": 15,  # Monto
            "G": 20,  # Técnico
        }

        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

        # Ajustar alineación
        for row in ws.iter_rows(min_row=row - len(historial_data["historial"]), max_row=row - 1):
            row[0].alignment = Alignment(horizontal="center")  # Fecha
            row[1].alignment = Alignment(horizontal="center")  # Documento
            row[2].alignment = Alignment(horizontal="center")  # Tipo
            row[4].alignment = Alignment(horizontal="center")  # Kilometraje
            row[5].alignment = Alignment(horizontal="right")  # Monto
            row[6].alignment = Alignment(horizontal="center")  # Técnico

        # Footer
        row += 2
        ws.merge_cells(f"A{row}:G{row}")
        ws[f"A{row}"] = (
            f"Documento generado por eGarage el {timezone.now().strftime('%d/%m/%Y %H:%M')}"
        )
        ws[f"A{row}"].font = Font(size=9, italic=True, color="666666")
        ws[f"A{row}"].alignment = Alignment(horizontal="center")

        # Guardar en BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Nombre del archivo
        patente_slug = vehiculo.patente.replace(" ", "_").replace("-", "_")
        filename = (
            f"Historial_Mantenimiento_{patente_slug}_{timezone.now().strftime('%Y%m%d')}.xlsx"
        )

        # Retornar Excel como respuesta
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    except ImportError:
        # Si openpyxl no está instalado, retornar error
        return HttpResponse(
            "Error: openpyxl no está instalado. Por favor, instálelo con: pip install openpyxl",
            status=500,
            content_type="text/plain",
        )
    except Exception as e:
        # Log del error y retornar mensaje
        import logging

        logger = logging.getLogger(__name__)
        logger.error(f"Error generando Excel de historial: {e}", exc_info=True)
        return HttpResponse(
            f"Error al generar Excel: {str(e)}", status=500, content_type="text/plain"
        )


@login_required_default
def verificar_garantia(request):
    """
    ✅ Vista para Verificar Trazabilidad de Garantías

    Permite verificar si un documento de garantía está dentro del límite.
    Puede recibir IDs de documentos o buscar automáticamente garantías activas.
    """
    # 🔒 FILTRO CRÍTICO POR EMPRESA
    empresa = get_user_empresa_safe(request.user)

    # Obtener IDs de documentos desde GET
    doc_garantia_id = request.GET.get("doc_garantia_id")
    doc_original_id = request.GET.get("doc_original_id")
    vehiculo_id = request.GET.get("vehiculo_id")

    resultado = None
    error = None
    doc_garantia = None
    doc_original = None
    garantias_activas = []

    # Si se proporciona vehículo, buscar garantías activas automáticamente
    if vehiculo_id:
        try:
            from taller.models.vehiculos import Vehiculo

            vehiculo = Vehiculo.objects.get(pk=vehiculo_id, empresa=empresa)

            # Buscar documentos recientes del vehículo que podrían ser garantías
            documentos_recientes = (
                Documento.objects.filter(
                    empresa=empresa, vehiculo=vehiculo, tipo__in=["OT", "PRES"], estado="EMITIDO"
                )
                .select_related("registro_kilometraje")
                .order_by("-fecha_emision")[:10]
            )

            # Buscar documentos anteriores que podrían ser el original
            for doc_reciente in documentos_recientes:
                try:
                    registro_reciente = doc_reciente.registro_kilometraje
                    if registro_reciente:
                        # Buscar documentos anteriores del mismo vehículo
                        docs_anteriores = (
                            Documento.objects.filter(
                                empresa=empresa,
                                vehiculo=vehiculo,
                                tipo__in=["OT", "PRES"],
                                fecha_emision__lt=doc_reciente.fecha_emision,
                                estado="EMITIDO",
                            )
                            .select_related("registro_kilometraje")
                            .order_by("-fecha_emision")[:5]
                        )

                        for doc_anterior in docs_anteriores:
                            try:
                                registro_anterior = doc_anterior.registro_kilometraje
                                if registro_anterior:
                                    reporte = ReporteKilometraje(empresa)
                                    verificacion = reporte.verificar_garantia(
                                        doc_reciente, doc_anterior
                                    )
                                    if verificacion and not verificacion.get("error"):
                                        garantias_activas.append(
                                            {
                                                "documento_garantia": doc_reciente,
                                                "documento_original": doc_anterior,
                                                "verificacion": verificacion,
                                            }
                                        )
                            except (ObjectDoesNotExist, AttributeError):
                                # Documento anterior no tiene registro_kilometraje asociado
                                continue
                except (ObjectDoesNotExist, AttributeError):
                    # Documento reciente no tiene registro_kilometraje asociado
                    continue
        except Vehiculo.DoesNotExist:
            error = "Vehículo no encontrado"
        except Exception as e:
            error = f"Error al buscar garantías: {str(e)}"

    # Si se proporcionan IDs específicos, verificar esa garantía
    if doc_garantia_id and doc_original_id:
        try:
            # Obtener documentos (con filtro de empresa)
            doc_garantia = Documento.objects.get(pk=doc_garantia_id, empresa=empresa)
            doc_original = Documento.objects.get(pk=doc_original_id, empresa=empresa)

            # Crear instancia del reporte
            reporte = ReporteKilometraje(empresa)

            # Verificar garantía
            resultado = reporte.verificar_garantia(doc_garantia, doc_original)

        except Documento.DoesNotExist:
            error = "Uno o ambos documentos no fueron encontrados"
        except Exception as e:
            error = f"Error al verificar garantía: {str(e)}"

    # Preparar contexto
    context = {
        "resultado": resultado,
        "error": error,
        "documento_garantia": doc_garantia,
        "documento_original": doc_original,
        "garantias_activas": garantias_activas,
        "vehiculo_id": vehiculo_id,
        "empresa": empresa,
    }

    return render(request, "taller/reportes/verificar_garantia.html", context)
