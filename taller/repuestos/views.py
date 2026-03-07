import json
import logging

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import models
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404
from django.template.loader import render_to_string
from django.views.decorators.http import require_GET
from django.views.decorators.csrf import csrf_exempt

from taller.models.empresa import Empresa
from taller.models.repuesto import Repuesto

from .views_cbv import (
    RepuestoCreateView,
    RepuestoDetailView,
    RepuestoListView,
    RepuestoUpdateView,
)

log = logging.getLogger(__name__)


def lista_repuestos(request, *args, **kwargs):
    log.info("FBV shim: lista_repuestos")
    return RepuestoListView.as_view()(request, *args, **kwargs)


def ver_repuesto(request, *args, **kwargs):
    log.info("FBV shim: ver_repuesto")
    return RepuestoDetailView.as_view()(request, *args, **kwargs)


def crear_repuesto(request, *args, **kwargs):
    log.info("FBV shim: crear_repuesto")
    return RepuestoCreateView.as_view()(request, *args, **kwargs)


def editar_repuesto(request, *args, **kwargs):
    log.info("FBV shim: editar_repuesto")
    return RepuestoUpdateView.as_view()(request, *args, **kwargs)


def eliminar_repuesto(request, pk):
    """Eliminar un repuesto - FILTRADO POR EMPRESA"""
    if request.method == "POST":
        try:
            # 🔒 BLINDAJE MULTI-TENANT: SIEMPRE filtrar por empresa
            empresa = getattr(request.user, "empresa", None)
            if not empresa:
                return JsonResponse(
                    {"success": False, "error": "Usuario sin empresa asignada"}, status=403
                )

            repuesto = get_object_or_404(Repuesto, pk=pk, empresa=empresa)
            repuesto.delete()
            messages.success(request, "Repuesto eliminado exitosamente.")
            return JsonResponse({"success": True})
        except Exception as e:
            log.error(f"Error al eliminar repuesto: {e}")
            return JsonResponse({"success": False, "error": str(e)}, status=500)
    else:
        return JsonResponse({"success": False, "error": "Método no permitido"}, status=405)


@login_required
@login_required
@csrf_exempt
def buscar_repuestos_ajax(request):
    """Búsqueda inteligente en tiempo real de repuestos"""
    try:
        log.info(f"Búsqueda AJAX iniciada - Método: {request.method}")
        log.info(f"Usuario: {request.user}")

        # Verificar que sea método POST
        if request.method != "POST":
            log.error(f"Método no permitido: {request.method}")
            return JsonResponse({"error": "Método no permitido"}, status=405)

        # Obtener empresa del usuario
        try:
            empresa = Empresa.objects.get(user=request.user)
            log.info(f"Empresa encontrada: {empresa}")
        except Empresa.DoesNotExist:
            empresa, created = Empresa.objects.get_or_create(
                user=request.user,
                defaults={"nombre_taller": f"Taller de {request.user.username}"},
            )
            log.info(f"Empresa {'creada' if created else 'encontrada'}: {empresa}")

        # Obtener query desde el JSON del request
        data = json.loads(request.body)
        query = data.get("query", "").strip()
        log.info(f"Query recibida: '{query}'")

        # Filtrar repuestos por empresa del usuario
        repuestos = Repuesto.objects.select_related("categoria").filter(empresa=empresa)
        log.info(f"Repuestos base encontrados: {repuestos.count()}")

        # Si hay query, filtrar por múltiples campos
        if query:
            repuestos = repuestos.filter(
                models.Q(nombre__icontains=query)
                | models.Q(part_number__icontains=query)
                | models.Q(categoria__nombre__icontains=query)
                | models.Q(proveedor__icontains=query)
            )
            log.info(f"Repuestos después del filtro: {repuestos.count()}")

        # Ordenar por nombre
        repuestos = repuestos.order_by("nombre")[:50]  # Limitar a 50 resultados

        # Obtener país para el contexto del template
        country = "CL"  # default
        if empresa and hasattr(empresa, "pais") and empresa.pais:
            country = empresa.pais
        elif hasattr(request, "country") and request.country:
            country = request.country
        elif request.path.startswith("/us/"):
            country = "US"
        elif request.path.startswith("/cl/"):
            country = "CL"

        # Renderizar template parcial
        html = render_to_string(
            "taller/common/repuestos/tabla_repuestos_ajax.html",
            {"repuestos": repuestos, "country": country, "request": request},
        )

        result = {"html": html, "total": repuestos.count()}
        log.info(f"Enviando respuesta con {result['total']} resultados")
        return JsonResponse(result)

    except Exception as e:
        log.error(f"Error en búsqueda AJAX: {str(e)}")
        import traceback

        log.error(f"Traceback: {traceback.format_exc()}")
        return JsonResponse({"error": "Error interno del servidor"}, status=500)


# --- Vistas AJAX de repuestos (repuesto_info, etc.) ---


@login_required
@require_GET
def repuesto_info(request):
    """
    AJAX: devuelve JSON con datos de un repuesto por id.
    Filtrado por empresa del usuario (multi-tenant).
    """
    try:
        repuesto_id = request.GET.get("id") or request.GET.get("pk")
        if not repuesto_id:
            return JsonResponse({"error": "Falta id o pk"}, status=400)

        empresa = getattr(request.user, "empresa", None)
        if not empresa:
            return JsonResponse({"error": "Usuario sin empresa"}, status=403)

        repuesto = get_object_or_404(Repuesto, pk=repuesto_id, empresa=empresa)
        return JsonResponse(
            {
                "id": repuesto.id,
                "part_number": repuesto.part_number or "",
                "nombre": repuesto.nombre or "",
                "precio_compra": str(repuesto.precio_compra or 0),
                "precio_venta": str(repuesto.precio_venta or 0),
                "cantidad_stock": repuesto.cantidad_stock or 0,
                "proveedor": repuesto.proveedor or "",
                "categoria": repuesto.categoria.nombre if repuesto.categoria else "",
            }
        )
    except ValueError:
        return JsonResponse({"error": "id inválido"}, status=400)


@login_required
@require_GET
def exportar_excel_repuestos(request):
    """
    Exporta el listado de repuestos de la empresa a Excel (.xlsx).
    Filtrado por empresa del usuario (multi-tenant).
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
    except ImportError:
        return JsonResponse(
            {"error": "openpyxl no instalado. Ejecuta: pip install openpyxl"},
            status=501,
        )

    empresa = getattr(request.user, "empresa", None)
    if not empresa:
        return JsonResponse({"error": "Usuario sin empresa"}, status=403)

    qs = Repuesto.objects.filter(empresa=empresa).select_related("categoria").order_by("nombre")
    wb = Workbook()
    ws = wb.active
    ws.title = "Repuestos"

    headers = ["Código", "Nombre", "Categoría", "P. compra", "P. venta", "Stock", "Proveedor"]
    font_bold = Font(bold=True)
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = font_bold
        c.alignment = Alignment(horizontal="center")

    for row, r in enumerate(qs, 2):
        ws.cell(row=row, column=1, value=r.part_number or "")
        ws.cell(row=row, column=2, value=r.nombre or "")
        ws.cell(row=row, column=3, value=r.categoria.nombre if r.categoria else "")
        ws.cell(row=row, column=4, value=float(r.precio_compra or 0))
        ws.cell(row=row, column=5, value=float(r.precio_venta or 0))
        ws.cell(row=row, column=6, value=r.cantidad_stock or 0)
        ws.cell(row=row, column=7, value=r.proveedor or "")

    from io import BytesIO
    from datetime import datetime

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = (
        f"repuestos_{empresa.nombre_taller or 'taller'}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    )
    filename = "".join(c if c.isalnum() or c in "._-" else "_" for c in filename)

    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp
