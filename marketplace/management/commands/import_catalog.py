"""
Comando de gestión para importar catálogos de proveedores desde Excel.
Permite precargar precios localmente como respaldo cuando las APIs externas fallan.

Uso:
    python manage.py import_catalog --casa "Indra" --file "catalogo_indra.xlsx"
    python manage.py import_catalog --casa "Indra" --file "catalogo_indra.xlsx" --update
"""
import os
from decimal import Decimal
from pathlib import Path

import openpyxl
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from marketplace.models import CasaRepuestos, ProductoCatalogo
from taller.models.empresa import Empresa


class Command(BaseCommand):
    help = "Importa catálogos de proveedores desde archivos Excel (.xlsx)"

    def add_arguments(self, parser):
        parser.add_argument(
            '--casa',
            type=str,
            required=True,
            help='Nombre de la casa de repuestos (ej: "Indra", "Bosch")'
        )
        parser.add_argument(
            '--file',
            type=str,
            required=True,
            help='Ruta al archivo Excel (.xlsx)'
        )
        parser.add_argument(
            '--empresa',
            type=int,
            help='ID de la empresa (opcional, si no se especifica se importa para todas)'
        )
        parser.add_argument(
            '--update',
            action='store_true',
            help='Actualizar productos existentes en lugar de solo crear nuevos'
        )
        parser.add_argument(
            '--sheet',
            type=str,
            default='Sheet1',
            help='Nombre de la hoja a importar (default: Sheet1)'
        )
        parser.add_argument(
            '--skip-rows',
            type=int,
            default=0,
            help='Número de filas a saltar al inicio (para encabezados, default: 0)'
        )

    def handle(self, *args, **options):
        casa_nombre = options['casa']
        file_path = options['file']
        empresa_id = options.get('empresa')
        update = options.get('update', False)
        sheet_name = options.get('sheet', 'Sheet1')
        skip_rows = options.get('skip_rows', 0)

        # Validar que el archivo existe
        if not os.path.exists(file_path):
            raise CommandError(f'El archivo "{file_path}" no existe.')

        if not file_path.endswith(('.xlsx', '.xls')):
            raise CommandError('El archivo debe ser Excel (.xlsx o .xls)')

        self.stdout.write(self.style.SUCCESS(f'📦 Importando catálogo de {casa_nombre}...'))
        self.stdout.write(f'   Archivo: {file_path}')

        try:
            # Cargar archivo Excel
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            if sheet_name not in workbook.sheetnames:
                raise CommandError(f'La hoja "{sheet_name}" no existe. Hojas disponibles: {", ".join(workbook.sheetnames)}')
            
            sheet = workbook[sheet_name]
            self.stdout.write(f'   Hoja: {sheet_name}')

            # Obtener empresas
            if empresa_id:
                empresas = [Empresa.objects.get(id=empresa_id)]
                self.stdout.write(f'   Empresa específica: {empresas[0].nombre_taller}')
            else:
                empresas = Empresa.objects.all()
                self.stdout.write(f'   Importando para todas las empresas ({empresas.count()} empresas)')

            total_creados = 0
            total_actualizados = 0
            total_errores = 0

            # Procesar cada empresa
            for empresa in empresas:
                self.stdout.write(f'\n   Procesando: {empresa.nombre_taller}')

                # Obtener o crear casa de repuestos
                casa, created = CasaRepuestos.objects.get_or_create(
                    empresa=empresa,
                    nombre=casa_nombre,
                    defaults={'activa': True}
                )
                if created:
                    self.stdout.write(self.style.SUCCESS(f'      ✅ Casa de repuestos "{casa_nombre}" creada'))
                else:
                    self.stdout.write(f'      ℹ️  Casa de repuestos "{casa_nombre}" ya existe')

                # Procesar filas del Excel
                # Formato esperado: part_number | nombre | precio_referencia | disponible (opcional)
                for idx, row in enumerate(sheet.iter_rows(min_row=1 + skip_rows, values_only=True), start=1 + skip_rows):
                    if not any(row):  # Fila vacía
                        continue

                    try:
                        # Intentar diferentes formatos comunes
                        part_number = str(row[0]).strip() if row[0] else None
                        nombre = str(row[1]).strip() if len(row) > 1 and row[1] else part_number or 'Sin nombre'
                        precio_str = row[2] if len(row) > 2 else '0'
                        disponible = bool(row[3]) if len(row) > 3 and row[3] is not None else True

                        # Validar part_number
                        if not part_number or part_number.lower() in ('none', 'null', ''):
                            continue

                        # Convertir precio
                        try:
                            if isinstance(precio_str, (int, float)):
                                precio_referencia = Decimal(str(precio_str))
                            else:
                                precio_str = str(precio_str).replace('$', '').replace(',', '').strip()
                                precio_referencia = Decimal(precio_str) if precio_str else Decimal('0.00')
                        except (ValueError, TypeError):
                            self.stdout.write(self.style.WARNING(f'      ⚠️  Fila {idx}: Precio inválido "{precio_str}", usando 0'))
                            precio_referencia = Decimal('0.00')

                        # Crear o actualizar producto
                        with transaction.atomic():
                            producto, created = ProductoCatalogo.objects.get_or_create(
                                empresa=empresa,
                                casa_repuestos=casa,
                                part_number=part_number,
                                defaults={
                                    'nombre': nombre,
                                    'precio_referencia': precio_referencia,
                                    'disponible': disponible,
                                    'activo': True,
                                }
                            )

                            if created:
                                total_creados += 1
                                if total_creados % 10 == 0:
                                    self.stdout.write(f'         ✅ {total_creados} productos creados...')
                            elif update:
                                producto.nombre = nombre
                                producto.precio_referencia = precio_referencia
                                producto.disponible = disponible
                                producto.activo = True
                                producto.save()
                                total_actualizados += 1
                                if total_actualizados % 10 == 0:
                                    self.stdout.write(f'         🔄 {total_actualizados} productos actualizados...')

                    except Exception as e:
                        total_errores += 1
                        self.stdout.write(self.style.ERROR(f'      ❌ Error en fila {idx}: {str(e)}'))
                        continue

            # Resumen
            self.stdout.write(self.style.SUCCESS(f'\n✅ Importación completada:'))
            self.stdout.write(f'   ✅ Productos creados: {total_creados}')
            if update:
                self.stdout.write(f'   🔄 Productos actualizados: {total_actualizados}')
            if total_errores > 0:
                self.stdout.write(self.style.WARNING(f'   ⚠️  Errores: {total_errores}'))

            self.stdout.write(self.style.SUCCESS('\n💡 Tip: Ejecuta este comando semanalmente para mantener el catálogo actualizado.'))

        except Exception as e:
            raise CommandError(f'Error al importar catálogo: {str(e)}')
