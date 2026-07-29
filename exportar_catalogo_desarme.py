"""
Script standalone: exporta CATALOGO_PIEZAS + PIEZAS_OPCIONALES a .xlsx para revisión de precios.
Sin Django. Ejecutar desde la raíz del proyecto:
    python exportar_catalogo_desarme.py
"""

import sys
import os
from pathlib import Path

# Permitir import relativo desde raíz sin instalar el paquete
sys.path.insert(0, str(Path(__file__).parent))

from taller.desarme.catalogo_piezas import CATALOGO_PIEZAS, PIEZAS_OPCIONALES
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

COLUMNAS = [
    "Código",
    "Nombre",
    "Zona",
    "Tipo",
    "Precio actual (CLP)",
    "Precio corregido",
    "¿Lo vendes? Sí/No",
    "Notas",
]

ANCHOS_MIN = [12, 36, 16, 12, 20, 18, 20, 30]


def build_rows():
    rows = []
    for codigo, nombre, zona, precio in CATALOGO_PIEZAS:
        rows.append((codigo, nombre, str(zona), "Estándar", int(precio), "", "", ""))
    for codigo, nombre, zona, precio in PIEZAS_OPCIONALES:
        rows.append((codigo, nombre, str(zona), "Opcional", int(precio), "", "", ""))
    rows.sort(key=lambda r: (r[2], r[0]))  # zona, luego código
    return rows


def autofit_columns(ws, anchos_min):
    for col_idx, col_cells in enumerate(ws.iter_cols(), start=1):
        max_len = anchos_min[col_idx - 1]
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)) + 2)
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len


def main():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Catálogo Desarme"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E79")
    center = Alignment(horizontal="center", vertical="center")

    for col_idx, col_name in enumerate(COLUMNAS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    ws.freeze_panes = "A2"

    rows = build_rows()
    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    autofit_columns(ws, ANCHOS_MIN)

    output_path = Path(__file__).parent / "catalogo_desarme_para_revision.xlsx"
    wb.save(output_path)

    total_filas = len(rows)
    print(f"Exportado: {output_path}")
    print(f"Total de filas de datos: {total_filas}  (esperado: 109)")


if __name__ == "__main__":
    main()
